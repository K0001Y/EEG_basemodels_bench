from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import tempfile
from pathlib import Path
from typing import Any

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REFERENCE = REPO_ROOT / "experiment_tracking" / "experiment_reference.csv"
DEFAULT_WORKBOOK = REPO_ROOT / "EEG实验进度追踪.xlsx"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "experiment_tracking"

MODEL_10 = [
    "reve",
    "cbramod",
    "biot",
    "brainomni",
    "femba",
    "neurogpt",
    "labram",
    "eegmamba",
    "neurolm",
    "bendr",
]

EXPERIMENT_SPEC = {
    "linear_full": {
        "columns": {"cross": "linear_full_cross", "within": "linear_full_within"},
        "seeds": [42, 10, 5],
        "models": MODEL_10,
        "result_root": "result_5/linear_full",
        "seed_dir": "seed_{seed}",
        "ratio_dir": "ratio_full",
        "dir_suffixes": ["linear_prob", "linear_full"],
    },
    "linear_downsample": {
        "columns": {"cross": "linear_downsample_cross", "within": "linear_downsample_within"},
        "seeds": [42, 10, 5],
        "models": MODEL_10,
        "result_root": "result_5/linear_downsample",
        "seed_dir": "seed_{seed}_downsample_t40",
        "ratio_dir": "ratio_full",
        "dir_suffixes": ["linear_downsample", "linear_prob"],
    },
    "full_finetune": {
        "columns": {"cross": "full_finetune_cross", "within": "full_finetune_within"},
        "seeds": [42, 10, 5],
        "models": MODEL_10,
        "result_roots": ["result_5/full_finetune", "result_5/full_finetune_alldata"],
        "seed_dirs": ["seed_{seed}", "seed_{seed}_downsample_t40"],
        "ratio_dir": "ratio_full",
        "dir_suffixes": ["full_finetune", "full_finetune_alldata"],
    },
    "fewshot": {
        "columns": {"cross": "fewshot_cross"},
        "seeds": [42, 10, 5, 2024, 3407],
        "models": MODEL_10,
        "result_root": "result_5/fewshot",
        "seed_dir": "seed_{seed}",
        "ratio_dirs": ["ratio_0p02", "ratio_0p05", "ratio_0p1", "ratio_0p3"],
        "dir_suffixes": ["fewshot"],
    },
    "channel_mask": {
        "columns": {"cross": "channel_mask_cross"},
        "seeds": [42, 10, 5, 2024, 3407],
        "models": MODEL_10,
        "result_root": "result_5/channel_mask",
        "seed_dir": "seed_{seed}_downsample_t40",
        "ratio_dirs": ["mask_0p2/ratio_full", "mask_0p4/ratio_full", "mask_0p6/ratio_full", "mask_0p8/ratio_full"],
        "dir_suffixes": ["channel_mask"],
    },
    "downsample_target_num": {
        "columns": {"cross": "downsample_target_num_cross"},
        "seeds": [42, 10, 5, 2024, 3407, 2025, 1234, 2022, 2023, 2026],
        "models": MODEL_10,
        "result_root": "downsample_sweeps",
        "target_nums": [5, 10, 20, 40, 80, 120],
        "records_file": "records.csv",
    },
}

HEADER_TO_KEY = {
    ("linear_downsample", "cross"): "linear_downsample_cross",
    ("linear_downsample", "within"): "linear_downsample_within",
    ("linear_full", "cross"): "linear_full_cross",
    ("linear_full", "within"): "linear_full_within",
    ("full_fintuen", None): "full_finetune_cross",
    ("full_finetune", None): "full_finetune_cross",
    ("downsample", None): "downsample_target_num_cross",
    ("channel\nmask ⭐", None): "channel_mask_cross",
    ("channel mask ⭐", None): "channel_mask_cross",
    ("few\nshot ⭐", None): "fewshot_cross",
    ("few shot ⭐", None): "fewshot_cross",
}


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_na(value: Any) -> bool:
    return _clean(value).upper() == "N/A"


def _normalize_status_cell(value: Any) -> str:
    return _clean(value)


def _normalize_dataset_key(value: str) -> str:
    key = _clean(value).lower()
    replacements = {
        "–": "_",
        "—": "_",
        "-": "_",
        " ": "_",
        "（": "(",
        "）": ")",
        "monitering": "monitoring",
        "extraversial": "extraversion",
        "fintuen": "finetune",
    }
    for src, dst in replacements.items():
        key = key.replace(src, dst)
    key = re.sub(r"\([^)]*\)", "", key)
    key = re.sub(r"_balanced$", "", key)
    key = re.sub(r"_old_badscale$", "", key)
    key = re.sub(r"_wsn$", "", key)
    return re.sub(r"[^a-z0-9]+", "", key)


def _dataset_matches(query: str, candidate: str) -> bool:
    q = _normalize_dataset_key(query)
    c = _normalize_dataset_key(candidate)
    if not q or not c:
        return False
    if q == c:
        return True
    if len(q) >= 5 and q in c:
        return True
    if len(c) >= 5 and c in q:
        return True
    return False


def _dataset_from_exp_dir_name(name: str, mode: str) -> str:
    marker = f"_{mode}_"
    if marker in name:
        return name.split(marker, 1)[0]
    return name


def _csv_task_columns() -> dict[tuple[str, str], str]:
    return {
        ("linear_downsample", "cross"): "linear_downsample_cross",
        ("linear_downsample", "within"): "linear_downsample_within",
        ("linear_full", "cross"): "linear_full_cross",
        ("linear_full", "within"): "linear_full_within",
        ("full_finetune", "cross"): "full_finetune",
        ("downsample_target_num", "cross"): "downsample_target_num",
        ("channel_mask", "cross"): "channel_mask",
        ("fewshot", "cross"): "fewshot",
    }


def _parse_headers(ws) -> dict[str, int]:
    top = [cell.value for cell in ws[1]]
    sub = [cell.value for cell in ws[2]]
    out: dict[str, int] = {}
    current = None
    for idx, value in enumerate(top, start=1):
        if value is not None:
            current = _clean(value)
        sub_value = _clean(sub[idx - 1]) or None
        key = HEADER_TO_KEY.get((current, sub_value)) or HEADER_TO_KEY.get((current, None))
        if key:
            out[key] = idx
    return out


def _iter_dataset_rows(ws):
    for row_idx in range(3, ws.max_row + 1):
        dataset = _clean(ws.cell(row_idx, 2).value)
        zarr_name = _clean(ws.cell(row_idx, 3).value)
        if dataset and zarr_name:
            yield row_idx, dataset, zarr_name


def _candidate_dirs(result_root: Path, dataset: str, mode: str, suffixes: list[str]) -> list[Path]:
    if not result_root.exists():
        return []
    exact_prefixes = [f"{dataset}_{mode}_"]
    if not dataset.endswith("_wsn"):
        exact_prefixes.append(f"{dataset}_wsn_{mode}_")
    dirs = []
    for path in result_root.iterdir():
        if not path.is_dir():
            continue
        if any(path.name.startswith(prefix) for prefix in exact_prefixes):
            dirs.append(path)
            continue
        if f"_{mode}_" not in path.name:
            continue
        candidate_dataset = _dataset_from_exp_dir_name(path.name, mode)
        if _dataset_matches(dataset, candidate_dataset):
            dirs.append(path)

    def score(path: Path) -> tuple[int, int, str]:
        candidate_dataset = _dataset_from_exp_dir_name(path.name, mode)
        exact = 0 if candidate_dataset == dataset or candidate_dataset == f"{dataset}_wsn" else 1
        for i, suffix in enumerate(suffixes):
            if path.name.endswith(f"_{suffix}"):
                return (exact, i, path.name)
        return (exact, len(suffixes), path.name)
    return sorted(dirs, key=score)


def _result_roots(spec: dict[str, Any]) -> list[Path]:
    roots = spec.get("result_roots")
    if roots is None:
        roots = [spec["result_root"]]
    return [REPO_ROOT / str(root) for root in roots]


def _seed_dir_names(spec: dict[str, Any], seed: int) -> list[str]:
    templates = spec.get("seed_dirs")
    if templates is None:
        templates = [spec["seed_dir"]]
    return [str(template).format(seed=seed) for template in templates]


def _models_from_leaderboard(path: Path, expected_models: set[str]) -> set[str]:
    if not path.exists():
        return set()
    try:
        data = json.loads(path.read_text())
    except Exception:
        return set()
    return {
        str(row.get("model", "")).lower()
        for row in data
        if str(row.get("model", "")).lower() in expected_models
    }


def _models_in_result_dir(path: Path, expected_models: set[str]) -> set[str]:
    found = set()
    for model in expected_models:
        model_dir = path / model
        if (model_dir / "summary.json").exists() or (model_dir / "train_log.jsonl").exists():
            found.add(model)
    found |= _models_from_leaderboard(path / "leaderboard_test.json", expected_models)
    return found


def _scan_standard_task(dataset: str, mode: str, spec: dict[str, Any]) -> dict[str, Any]:
    expected_models = set(spec["models"])
    dirs = []
    for result_root in _result_roots(spec):
        dirs.extend(_candidate_dirs(result_root, dataset, mode, spec.get("dir_suffixes", [])))
    ratio_dirs = spec.get("ratio_dirs") or [spec.get("ratio_dir", "ratio_full")]
    seed_details = []
    complete_units = 0
    total_units = len(spec["seeds"]) * len(ratio_dirs)
    for seed in spec["seeds"]:
        for ratio_dir in ratio_dirs:
            best_models: set[str] = set()
            best_path = None
            for exp_dir in dirs:
                for seed_dir_name in _seed_dir_names(spec, seed):
                    path = exp_dir / seed_dir_name / ratio_dir
                    models = _models_in_result_dir(path, expected_models)
                    if len(models) > len(best_models):
                        best_models = models
                        best_path = path
            missing = sorted(expected_models - best_models)
            if not missing:
                complete_units += 1
            seed_details.append({
                "seed": seed,
                "unit": ratio_dir,
                "path": str(best_path.relative_to(REPO_ROOT)) if best_path else "",
                "completed_models": sorted(best_models),
                "completed_model_count": len(best_models),
                "missing_models": missing,
            })
    status = "complete" if complete_units == total_units else ("not_run" if all(d["completed_model_count"] == 0 for d in seed_details) else "partial")
    return {
        "status": status,
        "complete_units": complete_units,
        "total_units": total_units,
        "candidate_dirs": [str(p.relative_to(REPO_ROOT)) for p in dirs],
        "details": seed_details,
    }


def _scan_downsample_target_num(dataset: str, mode: str, spec: dict[str, Any]) -> dict[str, Any]:
    records = REPO_ROOT / spec["result_root"] / dataset / mode / "linear_probing" / spec["records_file"]
    expected = {
        (int(seed), str(model), int(target_num))
        for seed in spec["seeds"]
        for model in spec["models"]
        for target_num in spec["target_nums"]
    }
    found = set()
    if records.exists():
        with records.open(newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    seed = int(row.get("seed", ""))
                    model = str(row.get("model", "")).lower()
                    target_num = int(float(row.get("target_num", "")))
                except Exception:
                    continue
                if (seed, model, target_num) in expected:
                    found.add((seed, model, target_num))
    missing = sorted(expected - found)
    status = "complete" if len(found) == len(expected) else ("not_run" if not found else "partial")
    return {
        "status": status,
        "complete_units": len(found),
        "total_units": len(expected),
        "candidate_dirs": [str(records.relative_to(REPO_ROOT))] if records.exists() else [],
        "details": [
            {"seed": seed, "model": model, "target_num": target_num}
            for seed, model, target_num in missing[:2000]
        ],
    }


def _status_text(result: dict[str, Any]) -> str:
    if result["status"] == "complete":
        return f"DONE {result['complete_units']}/{result['total_units']}"
    if result["status"] == "not_run":
        return "NOT_RUN"
    return f"PARTIAL {result['complete_units']}/{result['total_units']}"


def build_dataset_map(wb, header_map: dict[str, int]) -> dict[str, dict[str, list[str]]]:
    ws = wb["实验进度追踪"]
    dataset_map = {task: {mode: [] for mode in spec["columns"]} for task, spec in EXPERIMENT_SPEC.items()}
    for row_idx, _, zarr_name in _iter_dataset_rows(ws):
        for task, spec in EXPERIMENT_SPEC.items():
            for mode, col_key in spec["columns"].items():
                col = header_map.get(col_key)
                if col is None:
                    continue
                dataset_map[task][mode].append(zarr_name)
    return dataset_map


def build_dataset_map_from_csv(reference: Path) -> tuple[dict[str, dict[str, list[str]]], list[dict[str, str]]]:
    dataset_map = {task: {mode: [] for mode in spec["columns"]} for task, spec in EXPERIMENT_SPEC.items()}
    with reference.open(newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    task_cols = _csv_task_columns()
    for row in rows:
        dataset = _clean(row.get("dataset") or row.get("数据集"))
        if not dataset:
            continue
        run_name = _clean(row.get("run_name") or row.get("zarr_path") or row.get("Zarr路径") or dataset)
        for (task, mode), col_name in task_cols.items():
            if mode not in EXPERIMENT_SPEC[task]["columns"]:
                continue
            dataset_map[task][mode].append(run_name)
    return dataset_map, rows


def write_updated_reference_csv(src: Path, rows: list[dict[str, Any]], out_path: Path) -> None:
    with src.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        ref_rows = list(reader)
        fieldnames = list(reader.fieldnames or [])
    task_cols = _csv_task_columns()
    status_by_dataset_task = {
        (_normalize_dataset_key(row["dataset"]), row["task"], row["mode"]): _status_text(row)
        for row in rows
    }
    for ref_row in ref_rows:
        dataset = _clean(ref_row.get("run_name") or ref_row.get("zarr_path") or ref_row.get("Zarr路径") or ref_row.get("dataset") or ref_row.get("数据集"))
        key_dataset = _normalize_dataset_key(dataset)
        for (task, mode), col_name in task_cols.items():
            if col_name not in fieldnames:
                continue
            status = status_by_dataset_task.get((key_dataset, task, mode))
            if status is not None:
                ref_row[col_name] = status
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(ref_rows)


def scan_all(dataset_map: dict[str, dict[str, list[str]]]) -> list[dict[str, Any]]:
    rows = []
    for task, by_mode in dataset_map.items():
        spec = EXPERIMENT_SPEC[task]
        for mode, datasets in by_mode.items():
            for dataset in datasets:
                if task == "downsample_target_num":
                    result = _scan_downsample_target_num(dataset, "cross_subject", spec)
                else:
                    mode_name = "cross_subject" if mode == "cross" else "within_subject"
                    result = _scan_standard_task(dataset, mode_name, spec)
                rows.append({
                    "task": task,
                    "mode": mode,
                    "dataset": dataset,
                    **result,
                })
    return rows


def write_dataset_lists(dataset_map: dict[str, dict[str, list[str]]], out_dir: Path) -> None:
    lists_dir = out_dir / "dataset_lists"
    lists_dir.mkdir(parents=True, exist_ok=True)
    for task, by_mode in dataset_map.items():
        for mode, datasets in by_mode.items():
            path = lists_dir / f"{task}_{mode}.txt"
            path.write_text(f"dataset_name: [{','.join(datasets)}]\n")


def write_summaries(rows: list[dict[str, Any]], out_dir: Path) -> None:
    summary_json = out_dir / "progress_summary.json"
    summary_json.write_text(json.dumps(rows, ensure_ascii=False, indent=2))

    summary_csv = out_dir / "progress_summary.csv"
    with summary_csv.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["task", "mode", "dataset", "status", "complete_units", "total_units"])
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row[k] for k in writer.fieldnames})

    missing_csv = out_dir / "missing_runs.csv"
    with missing_csv.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["task", "mode", "dataset", "seed", "unit", "missing_models", "status"])
        writer.writeheader()
        for row in rows:
            if row["status"] == "complete":
                continue
            for detail in row.get("details", []):
                missing_models = detail.get("missing_models")
                if missing_models is None:
                    missing_models = [detail.get("model", "")]
                writer.writerow({
                    "task": row["task"],
                    "mode": row["mode"],
                    "dataset": row["dataset"],
                    "seed": detail.get("seed", ""),
                    "unit": detail.get("unit", detail.get("target_num", "")),
                    "missing_models": ",".join(str(x) for x in missing_models),
                    "status": row["status"],
                })


def write_updated_workbook(src: Path, rows: list[dict[str, Any]], header_map: dict[str, int], out_path: Path) -> None:
    wb = openpyxl.load_workbook(src)
    ws = wb["实验进度追踪"]
    row_by_dataset = {zarr: row_idx for row_idx, _, zarr in _iter_dataset_rows(ws)}
    col_by_task_mode = {}
    for task, spec in EXPERIMENT_SPEC.items():
        for mode, col_key in spec["columns"].items():
            col_by_task_mode[(task, mode)] = header_map.get(col_key)

    for row in rows:
        row_idx = row_by_dataset.get(row["dataset"])
        col_idx = col_by_task_mode.get((row["task"], row["mode"]))
        if row_idx and col_idx:
            ws.cell(row_idx, col_idx).value = _status_text(row)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        shutil.copyfile(tmp_path, out_path)
    finally:
        tmp_path.unlink(missing_ok=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Update EEG experiment tracking files from results.")
    parser.add_argument("--reference", default=str(DEFAULT_REFERENCE), help="CSV or XLSX tracking reference table.")
    parser.add_argument("--workbook", default=None, help="Backward-compatible alias for --reference.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    args = parser.parse_args()

    reference = Path(args.workbook or args.reference)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    header_map = None
    if reference.suffix.lower() == ".csv":
        dataset_map, _ = build_dataset_map_from_csv(reference)
    else:
        wb_ro = openpyxl.load_workbook(reference, read_only=True, data_only=True)
        ws = wb_ro["实验进度追踪"]
        header_map = _parse_headers(ws)
        dataset_map = build_dataset_map(wb_ro, header_map)

    (out_dir / "experiment_spec.json").write_text(json.dumps(EXPERIMENT_SPEC, ensure_ascii=False, indent=2))
    (out_dir / "experiment_dataset_map.json").write_text(json.dumps(dataset_map, ensure_ascii=False, indent=2))
    write_dataset_lists(dataset_map, out_dir)

    rows = scan_all(dataset_map)
    write_summaries(rows, out_dir)
    if reference.suffix.lower() == ".csv":
        write_updated_reference_csv(reference, rows, out_dir / "experiment_reference_updated.csv")
    else:
        write_updated_workbook(reference, rows, header_map or {}, out_dir / "experiment_progress_updated.xlsx")

    counts = {}
    for row in rows:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
    print(json.dumps(counts, ensure_ascii=False, indent=2))
    print(f"Reference: {reference}")
    print(f"Wrote tracking outputs to: {out_dir}")


if __name__ == "__main__":
    main()
